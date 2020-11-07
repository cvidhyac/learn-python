# Simple Excel operations using openpyxl
# Load workbook using openpyxl.load_workbook
# Get Sheets using wrk_book.sheetnames::List
#

import openpyxl, os, logging


logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')


# Open a xlsx file and return the workbook object
def open_xls(directory_location, file_name) :
    os.chdir(directory_location)
    wrk_book = openpyxl.open(file_name)
    logging.info(wrk_book.sheetnames)
    return wrk_book


# Get sheet from workbook, and print sheet name
def print_cell_value(wrk_book, sheet_name):
    sheet = wrk_book[sheet_name]
    cell = sheet['A2']
    logging.info(cell.value)


def print_few_columns(work_book, sheet_name):
    sheet = work_book[sheet_name]
    for i in range(1, 10):
        print("The column value for " + "i=" +str(i) +  " is : ", sheet.cell(row=i, column=2).value)

def create_new_workbook():
  wb = openpyxl.Workbook()
  logging.info("The sheet names are : "+ str(wb.sheetnames))
  sheet = wb.create_sheet('data_sheet')
  sheet['A1'] = 'Title'
  sheet['A2'] = 'Data'
  sheet['B1'] = "Test"
  sheet['B2'] = " Test Data"
  sheet = wb['Sheet']
  sheet.title = 'Renamed'
  wb.save("test_workbook.xlsx")
  wb.close()


logging.info("Current directory is :" + os.getcwd())
wb = open_xls("./resources", "sample_data.xlsx")
print_cell_value(wb, 'Sheet1')
print_few_columns(wb, 'Sheet1')
create_new_workbook()

